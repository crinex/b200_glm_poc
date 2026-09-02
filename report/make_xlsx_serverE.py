#!/usr/bin/env python3
"""서버 E 구성별 벤치 결과 → Excel (results/b200_E_*.md 에서 직접 읽음)

TP=8·EP1·MTP1 / TP=8·EP1·MTP2 / TP=2·DP4·EP8·MTP1 비교.
4개 지표: TPOT / Interactivity / Input TPS / Output TPS.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
RESULTS = os.path.join(HERE, "..", "results")
OUT = os.path.join(HERE, "GLM5.2_서버E_구성비교.xlsx")

# (표시이름, md파일, 색상)
CONFIGS = [
    ("TP8·EP1·MTP1 (기준)", "b200_E_baseline_ref_sweep.md", "1F7DB4"),
    ("TP8·EP1·MTP2",        "b200_E_mtp2_sweep.md",         "D9822B"),
    ("TP2·DP4·EP8·MTP1",    "b200_E_dp4tp2_ep8_sweep.md",   "1B7F4D"),
]


def parse_md(path):
    """{conc: (tpot, inter, in_tps, out_tps)}"""
    rows = {}
    for l in open(path).read().split("\n"):
        if l.startswith("|8"):
            p = [x.strip() for x in l.split("|")]
            # |''|ISL|OSL|GPU|Prec|TP|p|d|conc|TPOT|Inter|InGPU|OutGPU|TotGPU|InSrv|OutSrv|
            rows[int(p[8])] = (float(p[9]), float(p[10]), int(p[14]), int(p[15]))
    return rows

DATA = [(name, parse_md(os.path.join(RESULTS, f)), color) for name, f, color in CONFIGS]
ALL_CONC = sorted({c for _, r, _ in DATA for c in r})

INK, DIM, HDR_BG, ZEBRA = "1A2230", "5B6A7D", "1F2937", "F4F7FA"
thin = Side(style="thin", color="D3DBE4")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
F_TITLE = Font(name="Malgun Gothic", size=14, bold=True, color=INK)
F_SUB = Font(name="Malgun Gothic", size=9.5, color=DIM)
F_HDR = Font(name="Malgun Gothic", size=9.5, bold=True, color="FFFFFF")
F_NUM = Font(name="Consolas", size=10, color=INK)
F_NOTE = Font(name="Malgun Gothic", size=9, color=DIM)
CEN = Alignment(horizontal="center", vertical="center")
RGT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

COND = ("gen8k 1,024장 · ISL≈8,200 · OSL 1,024 · Weight FP8 · KV fp8 · "
        "vLLM 0.28.0 · 서버 E (B200×8, 1000W, 드라이버 595)")


def put(ws, r, c, v, font=F_NUM, fmt=None, align=RGT, fill=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font, cell.alignment, cell.border = font, align, BORDER
    if fmt: cell.number_format = fmt
    if fill: cell.fill = PatternFill("solid", fgColor=fill)
    return cell


def hdr(ws, r, cols):
    for i, t in enumerate(cols, 1):
        c = ws.cell(row=r, column=i, value=t)
        c.font, c.alignment, c.border = F_HDR, WRAP, BORDER
        c.fill = PatternFill("solid", fgColor=HDR_BG)
    ws.row_dimensions[r].height = 30


wb = Workbook()
wb.remove(wb.active)

# ── 구성별 시트 ───────────────────────────────────────────────
COLS = ["conc", "TPOT (ms)", "Interactivity\n(tok/s/user)",
        "Input TPS\n/server", "Output TPS\n/server"]
for name, rows, color in DATA:
    sheet = name.split()[0].replace("·", "_")[:28]
    ws = wb.create_sheet(sheet)
    ws["A1"] = name
    ws["A1"].font = Font(name="Malgun Gothic", size=13, bold=True, color=color)
    ws.merge_cells("A1:E1")
    ws["A2"] = COND
    ws["A2"].font = F_SUB
    ws.merge_cells("A2:E2")
    hdr(ws, 4, COLS)
    r = 5
    for conc in ALL_CONC:
        if conc not in rows:
            continue
        tpot, inter, i_, o_ = rows[conc]
        z = ZEBRA if (r % 2 == 1) else None
        put(ws, r, 1, conc, align=CEN, fmt="0", fill=z)
        put(ws, r, 2, tpot, fmt="0.00", fill=z)
        put(ws, r, 3, inter, fmt="0.00", fill=z)
        put(ws, r, 4, i_, fmt="#,##0", fill=z)
        put(ws, r, 5, o_, fmt="#,##0", fill=z)
        r += 1
    for i, w in enumerate([8, 12, 15, 13, 13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

# ── 통합 비교 시트 (Output TPS 중심) ──────────────────────────
ws = wb.create_sheet("Output_비교", 0)
ws["A1"] = "Output TPS 통합 비교 (tok/s per server)"
ws["A1"].font = F_TITLE
ws.merge_cells("A1:F1")
ws["A2"] = COND
ws["A2"].font = F_SUB
ws.merge_cells("A2:F2")
cols = ["conc"] + [n for n, _, _ in DATA] + ["DP4 vs 기준"]
hdr(ws, 4, cols)
base = DATA[0][1]
r = 5
GREEN, RED = "1B7F4D", "B4341F"
for conc in ALL_CONC:
    z = ZEBRA if (r % 2 == 1) else None
    put(ws, r, 1, conc, align=CEN, fmt="0", fill=z)
    for j, (_, rows, _) in enumerate(DATA):
        v = rows.get(conc)
        put(ws, r, 2 + j, v[3] if v else "—",
            fmt="#,##0" if v else None, align=RGT if v else CEN, fill=z)
    # DP4 vs 기준
    b = base.get(conc); d = DATA[2][1].get(conc)
    if b and d:
        ratio = d[3] / b[3] - 1
        c = put(ws, r, 5, ratio, fmt="+0.0%;-0.0%", fill=z)
        c.font = Font(name="Consolas", size=10, bold=True,
                      color=GREEN if ratio >= 0 else RED)
    else:
        put(ws, r, 5, "—", align=CEN, fill=z)
    r += 1
r += 1
ws.cell(row=r, column=1, value=(
    "· TP2·DP4·EP8 은 KV cache 3.94M (기준 1.48M 의 2.67배). 고conc 에서 우세, 저conc 에서 열세.\n"
    "· 손익분기 conc=64. conc≥128 부터 역전. conc=256 에서 +47% (3,188 tok/s, 세션 최고).\n"
    "· TP8·EP1·MTP2 는 스크리닝이라 conc 32/128/256 만 측정.\n"
    "· TPOT/Interactivity 는 개별 사용자 체감 — DP4 여도 rank당 TP=2 라 기준과 유사. "
    "Output TPS(서버 총량)만 DP4 가 크게 이긴다.")).font = F_NOTE
ws.cell(row=r, column=1).alignment = Alignment(vertical="top", wrap_text=True)
ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=6)
for i, w in enumerate([8, 20, 16, 18, 13], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "B5"

wb.save(OUT)
print("저장:", os.path.abspath(OUT))
