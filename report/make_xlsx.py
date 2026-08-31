#!/usr/bin/env python3
"""GLM-5.2 가속기 벤치마크 비교표 → Excel"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "GLM5.2_벤치마크_비교.xlsx")

# ── 데이터 ────────────────────────────────────────────────────────────
# B200: 직접 측정. vLLM 0.28.0, TP=8, kv-cache-dtype fp8, mnbt 8192,
#       MTP 미적용, gen8k 1,024 고유 sheet, ISL ~8,200 / OSL 1,024
#       (OSL 1024 는 "8k1k 벤치" 원본 문서 조건 = --max-tokens 기본값)
B200 = [
    # conc, tpot, inter, inTPS, outTPS, totGPU
    (1,   9.45,  105.77, 815,   102,  115),
    (4,   11.48, 87.09,  2660,  333,  374),
    (8,   13.04, 76.72,  4614,  577,  649),
    (16,  15.97, 62.60,  7523,  940,  1058),
    (32,  20.70, 48.32,  11724, 1465, 1649),
    (64,  35.57, 28.11,  13883, 1734, 1952),
    (128, 62.69, 15.95,  15810, 1975, 2223),
    (256, 83.75, 11.94,  16138, 2016, 2269),
]
# H200 / Moreh: H200_GLM5.2_Measure.pdf
H200 = [
    (4,   59.06, 1627, 203),
    (8,   46.36, 2535, 317),
    (16,  31.96, 3578, 447),
    (32,  18.37, 4401, 550),
    (64,  11.27, 5520, 690),
    (128, 9.77,  5149, 643),
    (256, 9.82,  5227, 653),
]
MTP = [
    (4,   86.36, 2357, 294),
    (8,   54.75, 3209, 401),
    (16,  36.34, 4126, 515),
    (32,  21.93, 4999, 625),
    (64,  12.26, 5730, 715),
    (128, 11.24, 5570, 696),
    (256, 11.23, 5618, 702),
]
MOREH = [(256, 74.7, 13.39, 21996, 2777)]

# ── 스타일 ────────────────────────────────────────────────────────────
C_B200, C_H200, C_MTP, C_MOREH = "2F7DE0", "22A08A", "D9822B", "A545D8"
INK, DIM = "1A2230", "5B6A7D"
HDR_BG, ZEBRA = "1F2937", "F4F7FA"

thin = Side(style="thin", color="D3DBE4")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

F_TITLE = Font(name="Malgun Gothic", size=15, bold=True, color=INK)
F_SUB   = Font(name="Malgun Gothic", size=9.5, color=DIM)
F_HDR   = Font(name="Malgun Gothic", size=9.5, bold=True, color="FFFFFF")
F_DEV   = Font(name="Malgun Gothic", size=10, bold=True)
F_NUM   = Font(name="Consolas", size=10, color=INK)
F_NUMD  = Font(name="Consolas", size=10, color=DIM)
F_NOTE  = Font(name="Malgun Gothic", size=9, color=DIM)
F_SECT  = Font(name="Malgun Gothic", size=11, bold=True, color=INK)

CEN = Alignment(horizontal="center", vertical="center")
RGT = Alignment(horizontal="right", vertical="center")
LFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)


def hdr_row(ws, row, cols):
    for i, txt in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=txt)
        c.font = F_HDR
        c.fill = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = WRAP
        c.border = BORDER
    ws.row_dimensions[row].height = 32


def put(ws, row, col, val, *, font=F_NUM, fmt=None, align=RGT, fill=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font
    c.alignment = align
    c.border = BORDER
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    return c


# ══════════════════════════════════════════════════════════════════════
wb = Workbook()

# ── Sheet 1: 전체 측정값 ──────────────────────────────────────────────
ws = wb.active
ws.title = "전체 측정값"

ws["A1"] = "GLM-5.2 가속기 추론 벤치마크 비교"
ws["A1"].font = F_TITLE
ws.merge_cells("A1:G1")
ws.row_dimensions[1].height = 24

ws["A2"] = ("B200 측정 조건: ISL ≈ 8,200 tok · OSL 1,024 tok · gen8k 1,024 고유 sheet "
            "· 8 GPU/server · KV cache fp8 · mnbt 8192 · MTP 미적용")
ws["A2"].font = F_SUB
ws.merge_cells("A2:G2")

COLS = ["장비 / 설정", "Concurrency", "TPOT (ms)\n↓ 낮을수록 좋음",
        "Interactivity\n(tok/s/user) ↑", "Input TPS\n/server ↑",
        "Output TPS\n/server ↑", "Total TPS\n/GPU ↑"]
hdr_row(ws, 4, COLS)

r = 5
groups = [
    ("B200 (8) · vLLM TP=8 · MTP 미적용", C_B200,
     [(c, tp, it, i, o, t) for c, tp, it, i, o, t in B200]),
    ("H200 (8) · MTP 미사용", C_H200,
     [(c, 1000/it, it, i, o, None) for c, it, i, o in H200]),
    ("H200 (8) · MTP 사용", C_MTP,
     [(c, 1000/it, it, i, o, None) for c, it, i, o in MTP]),
    ("Moreh MI355X", C_MOREH,
     [(c, tp, it, i, o, None) for c, tp, it, i, o in MOREH]),
]

for gi, (label, color, rows) in enumerate(groups):
    start = r
    for j, (conc, tpot, inter, intps, outtps, totgpu) in enumerate(rows):
        z = ZEBRA if gi % 2 == 0 else None
        c1 = put(ws, r, 1, label if j == 0 else None,
                 font=F_DEV, align=LFT, fill=z)
        c1.font = Font(name="Malgun Gothic", size=10, bold=True, color=color)
        put(ws, r, 2, conc, font=F_NUM, align=CEN, fmt="0", fill=z)
        put(ws, r, 3, round(tpot, 2), fmt="0.00", fill=z)
        put(ws, r, 4, round(inter, 2), fmt="0.00", fill=z)
        put(ws, r, 5, intps, fmt="#,##0", fill=z)
        put(ws, r, 6, outtps, fmt="#,##0", fill=z)
        put(ws, r, 7, totgpu if totgpu is not None else "—",
            font=F_NUM if totgpu is not None else F_NUMD,
            fmt="#,##0" if totgpu is not None else None,
            align=RGT if totgpu is not None else CEN, fill=z)
        r += 1
    if len(rows) > 1:
        ws.merge_cells(start_row=start, start_column=1, end_row=r-1, end_column=1)
        ws.cell(row=start, column=1).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True)

r += 1
ws.cell(row=r, column=1, value=(
    "· B200은 vLLM 0.28.0 / TP=8 로 직접 측정. H200 및 Moreh 수치는 H200_GLM5.2_Measure.pdf 출처.\n"
    "· H200 계열 TPOT은 PDF 기재 Interactivity에서 역산 (1000 / Interactivity). PDF 값과 일치 확인.\n"
    "· Moreh는 PDF에 conc=256 단일 측정만 존재.\n"
    "· 주의: H200 / Moreh 의 서버 구성(mnbt, max-num-seqs, kv-cache-dtype, gpu-util, 실제 TP)은 "
    "문서화되어 있지 않아 확인 불가. ISL / OSL 도 PDF에 기재 없음.\n"
    "· 주의: Moreh precision 은 PDF 표에 FP8 로 표기되었으나, 원본 워크플로 문서의 예시 명령은 "
    "--precision MXFP4 이고 원본 스크립트 기본 모델 경로도 GLM-5.2-MXFP4 다. 상충 상태.\n"
    "· 따라서 이 표는 '측정된 처리량의 비교'이며 동일 조건 하드웨어 비교로 읽으면 안 된다."
)).font = F_NOTE
ws.cell(row=r, column=1).alignment = Alignment(vertical="top", wrap_text=True)
ws.merge_cells(start_row=r, start_column=1, end_row=r+6, end_column=7)

widths = [34, 13, 15, 15, 14, 14, 13]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"


# ── Sheet 2: 배수 비교 ────────────────────────────────────────────────
ws2 = wb.create_sheet("B200 배수 비교")

ws2["A1"] = "B200 대비 배수 — concurrency별"
ws2["A1"].font = F_TITLE
ws2.merge_cells("A1:H1")
ws2.row_dimensions[1].height = 24
ws2["A2"] = "1.00 초과 = B200 우위 / 1.00 미만 = 상대 장비 우위"
ws2["A2"].font = F_SUB
ws2.merge_cells("A2:H2")

b_by_c = {c: (tp, it, i, o) for c, tp, it, i, o, _ in B200}
h_by_c = {c: (it, i, o) for c, it, i, o in H200}
m_by_c = {c: (it, i, o) for c, it, i, o in MTP}

COLS2 = ["기준 장비", "Concurrency", "Interactivity\n배수", "Input TPS\n배수",
         "Output TPS\n배수", "B200 Interactivity", "B200 Input TPS", "B200 Output TPS"]
hdr_row(ws2, 4, COLS2)

GREEN, RED = "1B7F4D", "B4341F"
r = 5
for label, color, table in [("vs H200 MTP 미사용", C_H200, h_by_c),
                            ("vs H200 MTP 사용", C_MTP, m_by_c)]:
    start = r
    for conc in sorted(table):
        if conc not in b_by_c:
            continue
        bt, bi, bin_, bo = b_by_c[conc]
        hi, hin, ho = table[conc]
        c1 = ws2.cell(row=r, column=1, value=label if r == start else None)
        c1.font = Font(name="Malgun Gothic", size=10, bold=True, color=color)
        c1.alignment = LFT
        c1.border = BORDER
        put(ws2, r, 2, conc, align=CEN, fmt="0")
        for col, (bv, hv) in [(3, (bi, hi)), (4, (bin_, hin)), (5, (bo, ho))]:
            ratio = bv / hv
            c = put(ws2, r, col, round(ratio, 2), fmt='0.00"×"')
            c.font = Font(name="Consolas", size=10, bold=True,
                          color=GREEN if ratio >= 1 else RED)
        put(ws2, r, 6, round(bi, 2), font=F_NUMD, fmt="0.00")
        put(ws2, r, 7, bin_, font=F_NUMD, fmt="#,##0")
        put(ws2, r, 8, bo, font=F_NUMD, fmt="#,##0")
        r += 1
    ws2.merge_cells(start_row=start, start_column=1, end_row=r-1, end_column=1)
    ws2.cell(row=start, column=1).alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=True)

# Moreh (conc 256만)
bt, bi, bin_, bo = b_by_c[256]
mc, mtpot, mi, min_, mo = MOREH[0]
c1 = ws2.cell(row=r, column=1, value="vs Moreh MI355X")
c1.font = Font(name="Malgun Gothic", size=10, bold=True, color=C_MOREH)
c1.alignment = LFT
c1.border = BORDER
put(ws2, r, 2, 256, align=CEN, fmt="0")
for col, (bv, mv) in [(3, (bi, mi)), (4, (bin_, min_)), (5, (bo, mo))]:
    ratio = bv / mv
    c = put(ws2, r, col, round(ratio, 2), fmt='0.00"×"')
    c.font = Font(name="Consolas", size=10, bold=True,
                  color=GREEN if ratio >= 1 else RED)
put(ws2, r, 6, round(bi, 2), font=F_NUMD, fmt="0.00")
put(ws2, r, 7, bin_, font=F_NUMD, fmt="#,##0")
put(ws2, r, 8, bo, font=F_NUMD, fmt="#,##0")

r += 2
ws2.cell(row=r, column=1, value=(
    "· conc 32 이하: B200이 전 지표 우위 (Input 최대 3.8× / Output 1.9×).\n"
    "· conc 64 이상: B200 Interactivity 급락 → conc 256에서 H200 MTP·Moreh 모두에 열위.\n"
    "· conc 256 Output TPS: B200 1,265 vs Moreh 2,777 (Moreh 2.2× 우위).\n"
    "· B200 prefill 강세 / decode 약세 패턴. MTP 미적용 상태."
)).font = F_NOTE
ws2.cell(row=r, column=1).alignment = Alignment(vertical="top", wrap_text=True)
ws2.merge_cells(start_row=r, start_column=1, end_row=r+3, end_column=8)

for i, w in enumerate([22, 13, 15, 13, 14, 17, 15, 16], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A5"

wb.save(OUT)
print("저장:", OUT)
