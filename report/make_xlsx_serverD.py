#!/usr/bin/env python3
"""서버 D 3구성 실험 결과 → Excel (TPOT / Interactivity / Input / Output)"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "GLM5.2_서버D_3구성_비교.xlsx")

# conc, TPOT, Interactivity, Input TPS/server, Output TPS/server
EXP = [
    ("실험1 기준선 (EP1, MTP 없음)", "1F7DB4", [
        (4, 11.22, 89.15, 2471, 309), (8, 13.05, 76.66, 4575, 571),
        (16, 15.79, 63.35, 7419, 927), (32, 20.45, 48.91, 11288, 1410),
        (64, 32.33, 30.93, 14147, 1767), (128, 61.02, 16.39, 16034, 2003),
        (256, 83.88, 11.92, 16006, 1999)]),
    ("실험2 MTP spec=1 (EP1) ★최적", "1B7F4D", [
        (4, 8.13, 122.95, 3420, 427), (8, 9.69, 103.19, 6244, 780),
        (16, 12.00, 83.36, 9972, 1246), (32, 16.05, 62.31, 15002, 1874),
        (64, 29.71, 33.65, 16370, 2045), (128, 52.91, 18.90, 18019, 2251),
        (256, 74.13, 13.49, 17286, 2159)]),
    ("실험3 MTP spec=1 + EP8", "8A5AC2", [
        (4, 9.07, 110.23, 3226, 403), (8, 10.23, 97.77, 5885, 735),
        (16, 12.68, 78.89, 9473, 1183), (32, 17.10, 58.49, 14231, 1778),
        (64, 31.89, 31.36, 15327, 1915), (128, 55.20, 18.12, 17078, 2133),
        (256, 74.98, 13.34, 16633, 2078)]),
]

INK, DIM, HDR_BG, ZEBRA = "1A2230", "5B6A7D", "1F2937", "F4F7FA"
thin = Side(style="thin", color="D3DBE4")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
F_TITLE = Font(name="Malgun Gothic", size=14, bold=True, color=INK)
F_SUB = Font(name="Malgun Gothic", size=9.5, color=DIM)
F_HDR = Font(name="Malgun Gothic", size=9.5, bold=True, color="FFFFFF")
F_NUM = Font(name="Consolas", size=10, color=INK)
F_NOTE = Font(name="Malgun Gothic", size=9, color=DIM)
CEN = Alignment(horizontal="center", vertical="center")
RGT = Alignment(horizontal="right", vertical="center")
LFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

COLS = ["conc", "TPOT (ms)", "Interactivity\n(tok/s/user)",
        "Input TPS\n/server", "Output TPS\n/server"]
COND = ("gen8k 1,024장 · ISL ≈ 8,200 · OSL 1,024 고정 · sweep 4~256 · "
        "Weight FP8 · KV cache fp8 · TP=8 · DP=1 · vLLM 0.28.0 · "
        "서버 D (드라이버 595.91.07, 1000W)")


def put(ws, r, c, v, font=F_NUM, fmt=None, align=RGT, fill=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font, cell.alignment, cell.border = font, align, BORDER
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    return cell


def hdr(ws, r):
    for i, t in enumerate(COLS, 1):
        c = ws.cell(row=r, column=i, value=t)
        c.font, c.alignment, c.border = F_HDR, WRAP, BORDER
        c.fill = PatternFill("solid", fgColor=HDR_BG)
    ws.row_dimensions[r].height = 30


wb = Workbook()
wb.remove(wb.active)

# ── 실험별 시트 ───────────────────────────────────────────────
for title, color, rows in EXP:
    ws = wb.create_sheet(title.split()[0])
    ws["A1"] = title
    ws["A1"].font = Font(name="Malgun Gothic", size=13, bold=True, color=color)
    ws.merge_cells("A1:E1")
    ws["A2"] = COND
    ws["A2"].font = F_SUB
    ws.merge_cells("A2:E2")
    hdr(ws, 4)
    for j, (c, tpot, inter, i_, o_) in enumerate(rows):
        r = 5 + j
        z = ZEBRA if j % 2 == 0 else None
        put(ws, r, 1, c, align=CEN, fmt="0", fill=z)
        put(ws, r, 2, tpot, fmt="0.00", fill=z)
        put(ws, r, 3, inter, fmt="0.00", fill=z)
        put(ws, r, 4, i_, fmt="#,##0", fill=z)
        put(ws, r, 5, o_, fmt="#,##0", fill=z)
    for i, w in enumerate([8, 12, 15, 13, 13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

# ── 분해 시트 ─────────────────────────────────────────────────
ws = wb.create_sheet("분해")
ws["A1"] = "Output TPS 분해 — MTP 몫 · EP8 몫"
ws["A1"].font = F_TITLE
ws.merge_cells("A1:F1")
ws["A2"] = "MTP 몫 = 실험2 − 실험1,  EP8 몫 = 실험3 − 실험2 (동일 서버)"
ws["A2"].font = F_SUB
ws.merge_cells("A2:F2")
cols2 = ["conc", "실험1\n기준선", "실험2\nMTP1+EP1", "실험3\nMTP1+EP8",
         "MTP 몫", "EP8 몫"]
for i, t in enumerate(cols2, 1):
    c = ws.cell(row=4, column=i, value=t)
    c.font, c.alignment, c.border = F_HDR, WRAP, BORDER
    c.fill = PatternFill("solid", fgColor=HDR_BG)
ws.row_dimensions[4].height = 30
GREEN, RED = "1B7F4D", "B4341F"
for j in range(7):
    r = 5 + j
    conc = EXP[0][2][j][0]
    b, m, e = EXP[0][2][j][4], EXP[1][2][j][4], EXP[2][2][j][4]
    z = ZEBRA if j % 2 == 0 else None
    put(ws, r, 1, conc, align=CEN, fmt="0", fill=z)
    put(ws, r, 2, b, fmt="#,##0", fill=z)
    put(ws, r, 3, m, fmt="#,##0", fill=z)
    put(ws, r, 4, e, fmt="#,##0", fill=z)
    c5 = put(ws, r, 5, (m - b) / b, fmt="+0.0%;-0.0%", fill=z)
    c5.font = Font(name="Consolas", size=10, bold=True,
                   color=GREEN if m >= b else RED)
    c6 = put(ws, r, 6, (e - m) / m, fmt="+0.0%;-0.0%", fill=z)
    c6.font = Font(name="Consolas", size=10, bold=True,
                   color=GREEN if e >= m else RED)
r = 13
ws.cell(row=r, column=1, value=(
    "· MTP spec=1: 전 구간 이득 (+8~38%). draft 수락률 75.2%, accept length 1.75\n"
    "· EP8: 전 구간 -4~-6% 균일 열세 → 이 워크로드에서 미사용 권장\n"
    "· 최고 처리량: 실험2, conc=128 에서 2,251 tok/s")).font = F_NOTE
ws.cell(row=r, column=1).alignment = Alignment(vertical="top", wrap_text=True)
ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=6)
for i, w in enumerate([8, 12, 13, 13, 11, 11], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

wb.save(OUT)
print("저장:", OUT)
